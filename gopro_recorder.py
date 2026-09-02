import requests
import time
import os
import winsound
import threading
from openpyxl import load_workbook
from datetime import date

# ── Config ────────────────────────────────────────────────
CAM_A        = "http://172.26.191.51:8080"   # Camera A — Nurse
CAM_B        = "http://172.22.107.51:8080"   # Camera B — Patient
EXCEL_PATH   = "DataCollection_Checklist.xlsx"
BASE_FOLDER  = "E:\\VR_Dataset"

SUBJECTS = ["Suhaib", "Wesam", "Nour Majah", "Noor Ejeilat"]
ACTIONS  = {"Pain": "P", "Dizziness": "D", "Medication": "M"}
ACCURACY = {"Standard": "S", "Fault": "F"}

# ── Save Path Helper ──────────────────────────────────────
def get_save_folder(trial_code):
    """Build save path from trial code e.g. P1_P_S_T1 → E:\VR_Dataset\P1\Pain\Standard\T1"""
    parts    = trial_code.split("_")
    pair     = parts[0]
    action   = {"P": "Pain", "D": "Dizziness", "M": "Medication"}[parts[1]]
    accuracy = {"S": "Standard", "F": "Fault"}[parts[2]]
    trial    = parts[3]  # e.g. T1
    return os.path.join(BASE_FOLDER, pair, action, accuracy, trial)

def delete_recording(trial_code):
    """Delete the two MP4 files for a trial and clear its Excel row.

    Returns a dict with keys '_A' and '_B', each being one of:
      'deleted'   — file existed and was removed
      'missing'   — file was not found (already gone or never saved)
      'error:<msg>' — file existed but could not be deleted
    """
    folder = get_save_folder(trial_code)
    report = {}
    for suffix in ["_A", "_B"]:
        file_path = os.path.join(folder, f"{trial_code}{suffix}.MP4")
        if not os.path.exists(file_path):
            report[suffix] = "missing"
        else:
            try:
                os.remove(file_path)
                report[suffix] = "deleted"
            except Exception as e:
                report[suffix] = f"error:{e}"
    # Clear Excel row
    reset_trial(trial_code)
    return report

# ── Camera Helpers ────────────────────────────────────────
def beep(times=1):
    for _ in range(times):
        winsound.Beep(1000, 500)
        time.sleep(0.2)

def setup_camera(ip):
    # Required steps — failure aborts setup
    required_steps = [
        ("Enable wired USB control", f"{ip}/gopro/camera/control/wired_usb?p=1", 1),
        ("Load preset",              f"{ip}/gopro/camera/presets/load?id=655360",  2),
    ]
    for step_name, url, delay in required_steps:
        try:
            r = requests.get(url, timeout=5)
            if r.status_code != 200:
                raise RuntimeError(f"HTTP {r.status_code}")
        except Exception as e:
            raise RuntimeError(f"Setup step '{step_name}' failed on {ip}: {e}")
        time.sleep(delay)

    # Optional steps — log warning but do not abort
    optional_steps = [
        ("Set codec H.264", f"{ip}/gopro/camera/setting?setting=108&option=8", 1),
    ]
    for step_name, url, delay in optional_steps:
        try:
            r = requests.get(url, timeout=5)
            if r.status_code != 200:
                print(f"  [WARNING] Optional setup step '{step_name}' returned HTTP {r.status_code} on {ip} — skipping")
        except Exception as e:
            print(f"  [WARNING] Optional setup step '{step_name}' failed on {ip}: {e} — skipping")
        time.sleep(delay)

def setup_both():
    """Setup both cameras simultaneously. Raises RuntimeError if either fails."""
    errors = {}
    def _setup(ip, key):
        try:
            setup_camera(ip)
        except Exception as e:
            errors[key] = e
    threads = [
        threading.Thread(target=_setup, args=(CAM_A, "a")),
        threading.Thread(target=_setup, args=(CAM_B, "b")),
    ]
    for t in threads: t.start()
    for t in threads: t.join()
    if errors:
        msgs = " | ".join(str(e) for e in errors.values())
        raise RuntimeError(msgs)

def start_recording(ip):
    r = requests.get(f"{ip}/gopro/camera/shutter/start", timeout=5)
    if r.status_code != 200:
        time.sleep(2)
        r = requests.get(f"{ip}/gopro/camera/shutter/start", timeout=5)
    return r.status_code == 200

def start_both():
    """Start both cameras simultaneously. Returns (ok_a, ok_b)."""
    results = {}
    def do_start(ip, key):
        results[key] = start_recording(ip)
    threads = [
        threading.Thread(target=do_start, args=(CAM_A, "a")),
        threading.Thread(target=do_start, args=(CAM_B, "b")),
    ]
    for t in threads: t.start()
    for t in threads: t.join()
    return results.get("a", False), results.get("b", False)

def stop_recording(ip, retries=3, delay=2):
    for attempt in range(retries):
        try:
            r = requests.get(f"{ip}/gopro/camera/shutter/stop", timeout=5)
            if r.status_code == 200:
                return True
        except Exception:
            pass
        time.sleep(delay)
    return False

def stop_both():
    """Stop both cameras simultaneously."""
    threads = [
        threading.Thread(target=stop_recording, args=(CAM_A,)),
        threading.Thread(target=stop_recording, args=(CAM_B,)),
    ]
    for t in threads: t.start()
    for t in threads: t.join()

def get_last_file(ip, retries=5, delay=2):
    """Fetch the last recorded file, retrying until the camera flushes it."""
    for attempt in range(retries):
        try:
            r = requests.get(f"{ip}/gopro/media/list", timeout=5)
            data = r.json()
            if "media" in data and data["media"] and data["media"][-1]["fs"]:
                last_folder = data["media"][-1]["d"]
                last_file   = data["media"][-1]["fs"][-1]["n"]
                return last_folder, last_file
        except Exception:
            pass
        time.sleep(delay)
    raise RuntimeError(f"Camera at {ip} did not expose a media file after {retries} attempts")

def download_file(ip, folder, filename, save_as, log_fn=None):
    """Download a file and return (actual_bytes, expected_bytes).
    actual_bytes is the size written to disk; expected_bytes is from Content-Length (0 if unknown).
    """
    url = f"{ip}/videos/DCIM/{folder}/{filename}"
    r = requests.get(url, stream=True, timeout=120)
    total = int(r.headers.get("Content-Length", 0))
    os.makedirs(os.path.dirname(save_as), exist_ok=True)

    # Report every 10% of the file or every 100 MB, whichever is smaller
    report_every = min(100 * 1024 * 1024, total // 10) if total else 100 * 1024 * 1024
    downloaded = 0
    last_reported = -report_every  # trigger a report at 0% immediately

    with open(save_as, "wb") as f:
        for chunk in r.iter_content(chunk_size=1024 * 1024):
            f.write(chunk)
            downloaded += len(chunk)
            if log_fn and (downloaded - last_reported) >= report_every:
                if total:
                    pct = downloaded * 100 // total
                    log_fn(f"    {downloaded // (1024*1024)} MB / {total // (1024*1024)} MB ({pct}%)")
                else:
                    log_fn(f"    {downloaded // (1024*1024)} MB downloaded...")
                last_reported = downloaded

    actual = os.path.getsize(save_as)
    if log_fn and total:
        if actual == total:
            log_fn(f"    {total // (1024*1024)} MB / {total // (1024*1024)} MB (100%) ✓")
        else:
            log_fn(f"    ⚠ Size mismatch: got {actual // (1024*1024)} MB, expected {total // (1024*1024)} MB")
    return actual, total

def download_both(folder_a, file_a, save_a, folder_b, file_b, save_b, log_fn_a=None, log_fn_b=None):
    """Download from both cameras simultaneously.
    Returns (err_a, err_b, size_a, size_b) where size_x is (actual, expected) or None on error.
    """
    errors = {}
    sizes  = {}
    def do_download(ip, folder, filename, save_as, key, log_fn):
        try:
            sizes[key] = download_file(ip, folder, filename, save_as, log_fn=log_fn)
        except Exception as e:
            errors[key] = e
    threads = [
        threading.Thread(target=do_download, args=(CAM_A, folder_a, file_a, save_a, "a", log_fn_a)),
        threading.Thread(target=do_download, args=(CAM_B, folder_b, file_b, save_b, "b", log_fn_b)),
    ]
    for t in threads: t.start()
    for t in threads: t.join()
    return errors.get("a"), errors.get("b"), sizes.get("a"), sizes.get("b")

def delete_file(ip, folder, filename, retries=3, delay=2):
    url = f"{ip}/gopro/media/delete/file?path={folder}/{filename}"
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=5)
            if r.status_code == 200:
                return True
        except Exception:
            pass
        time.sleep(delay)
    return False

def wait_for_camera(ip, label, log_fn, retries=5, delay=3):
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(f"{ip}/gopro/camera/info", timeout=3)
            if r.status_code == 200:
                return True
        except Exception:
            pass
        log_fn(f"  ⚠ {label} not responding (attempt {attempt}/{retries})...")
        time.sleep(delay)
    return False

# ── Excel Helpers ─────────────────────────────────────────
def _get_trial_row(trial_code):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Master Checklist"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == trial_code:
            return wb, ws, row
    return wb, ws, None

def reset_trial(trial_code):
    """Clear Done, Date and Notes columns for a trial."""
    wb, ws, row = _get_trial_row(trial_code)
    if row:
        row[9].value  = None
        row[10].value = None
        row[11].value = None
        wb.save(EXCEL_PATH)
        return True
    return False

def mark_done(trial_code, remarks=None):
    wb, ws, row = _get_trial_row(trial_code)
    if row:
        row[9].value  = "✓"
        row[10].value = date.today().strftime("%Y-%m-%d")
        if remarks:
            existing = row[11].value or ""
            row[11].value = (existing + " | " + remarks).strip(" | ") if existing else remarks
        wb.save(EXCEL_PATH)
        return True
    return False

def add_remark(trial_code, remark):
    """Append a remark to the Notes column without marking done."""
    wb, ws, row = _get_trial_row(trial_code)
    if row:
        existing = row[11].value or ""
        row[11].value = (existing + " | " + remark).strip(" | ") if existing else remark
        wb.save(EXCEL_PATH)

def is_done(trial_code):
    _, _, row = _get_trial_row(trial_code)
    return row is not None and row[9].value == "✓"

def lookup_pair(nurse, patient):
    """Return pair number string (e.g. 'P1') by looking up nurse+patient in Pairs Reference sheet."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Pairs Reference"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).startswith("P"):
            pair, subj, _, subj_role, partner, _, partner_role = row[:7]
            if subj_role == "Nurse" and partner_role == "Patient":
                if subj == nurse and partner == patient:
                    return str(pair)
    return None